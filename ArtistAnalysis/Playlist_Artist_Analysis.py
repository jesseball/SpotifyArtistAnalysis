import sys
import spotipy
import spotipy.util as util
import json
import csv
import openpyxl
import collections

# Spotify codes, which will be set by each person. Please edit here with your specific codes.
ClientID 		= 'yourClientID'
ClientSecret 	= 'yourClientSecret'
RedirectURI 	= 'yourRedirectURI'
UserID 			= 'yourUserID'


# Set global variables. uri, uri2, and uri3 are simply example artist URI's to use for testing. They are contained in an array which can later be appended to add 
#	artists from a playlist.
# title is an array which gives column headers to the excel document which will be made in the process
scope 				= 'user-library-read'
uri 				= 'spotify:artist:74XFHRwlV6OrjEM0A2NCMF'
uri2				= 'spotify:artist:13y7CgLHjMVRMDqxdx0Xdo'
uri3				= 'spotify:artist:0Y5tJX1MQlPlqiwlOH1tJY'
wb 					= openpyxl.load_workbook('test2.xlsx')
sheet				= wb.get_sheet_by_name('Sheet1')
title 				= [("Artist"), ("SongTitle"), ("Danceability"), ("Energy"), ("Key"), ("Loudness"), ("Mode"), ("Speechiness"), 
						("Acousticness"), ("Instrumentalness"), ("Liveness"), ("Valence"), ("Tempo"), ("Type"), ("ID"), ("URI"), 
						("Track_Href"), ("analysis_url"), ("Duration (ms)"), ("Time Signature"), ('Genre')]
uris 				= []
genre 				= []
foundPlaylistURI 	= 'blank'
playlistNameToFind	= 'yourPlaylistName'


# This code uses the following functions from the Spotipy Library...
#	artist_top_tracks: receives a list of top 10 tracks from the artist (specified with URI)
#	audio_features: pulls an array of audio features from a specific song (specified with URI)
# It takes this information and it searches an artists top songs, and adds that information to an excel sheet
def artistTopSongSearch( artistURI , artistCount, sp ):

	TopSongs 			= sp.artist_top_tracks(artistURI, country='US')
	artistName 			= TopSongs['tracks'][0]['artists'][0]['name']
	song_artist_list 	= []
	RowIndex 			= 2 + (10*artistCount)
	ColumnIndex 		= 1

	getGenreFromArtist( artistURI, sp )
	genreString 	= " ".join(genre)
	counter 		= collections.Counter(genreString.split(" "))
	genreTuple		= counter.most_common()
	genreTermList	= list(genreTuple[0])
	genreMostCommon	= genreTermList[0]


	# This for loop iterates through each of the 10 top songs for an artist. For each one, it receives the audio features, adds them to an excel sheet
	# At the end it adds to the artistCount index, to ensure proper placement in the excel sheet
	for songIndex in TopSongs['tracks']:
		AudioFeatures = sp.audio_features(songIndex['uri'])
		sheet.cell(row=RowIndex, column = ColumnIndex).value = artistName
		ColumnIndex+=1
		sheet.cell(row=RowIndex, column = ColumnIndex).value = songIndex['name']
		ColumnIndex+=1
		for featuresIndex in AudioFeatures[0]:
			sheet.cell(row=RowIndex, column = ColumnIndex).value = AudioFeatures[0][featuresIndex]
			ColumnIndex+=1
		sheet.cell(row=RowIndex, column = ColumnIndex).value = genreMostCommon
		RowIndex+=1
		ColumnIndex=1
	
	artistCount = artistCount + 1

	del genre[:]


# This function receives "sp" and uses it to get the users lists of playlist. It does this with the Spotify function user_playlists
# This function then selects the first playlist, and for each of the songs in the playlist, adds that artists URI to the uri array global variables
# This is primarily for data collection, and is used to avoid manually adding ~20 artists to the uri array. Rather, a single song of each can be added to a playlist
# and that playlist is scanned for each artist
# Will add functionality to not add artists twice, and to strip individual songs from a playlist instead of just artists
def getArtistsFromPlaylist( sp ):
	global playlistNameToFind
	playlists = sp.user_playlists(user = UserID)
	findPlaylist( playlists, sp, playlistNameToFind )
	selectedPlaylistTracks = sp.user_playlist_tracks(user = UserID, playlist_id = foundPlaylistURI, limit = 30)
	for i in selectedPlaylistTracks['items']:
		#print(i['track']['name'] + " - " + i['track']['artists'][0]['name'])
		uris.append(i['track']['artists'][0]['uri'])


# This function pulls the genres from the Artist and puts them in an array
def getGenreFromArtist( ArtistURI, sp ):
	ArtistInfo	= sp.artist(ArtistURI)
	for genres in ArtistInfo['genres']:
		genre.append(genres)


# This function searches for a playlist in the playlist array based on a name provided
# It goes through the users playlists which are inputted, and iterates through to find a match
# This match is stored in the foundPlaylistURI field which is returned for later use
def findPlaylist( playlists, sp, playlistName ):
	global foundPlaylistURI
	for playlistList in playlists['items']:
		if (playlistList['name'] == playlistName):
			foundPlaylistURI = playlistList['uri']
	#print(foundPlaylistURI)
	return foundPlaylistURI



# In this function, the main code verifies the user with a token. Then the token verifies wtih Spotify, and the specific user. From there, it will select
# the users first playlist, and pull the artists from it into an array. For each index in that array, it will use the artistTopSongSearch() function to receive and analyze
# the top 10 songs from each artist. 
# It finalizes by saving the workbook.
def main():
	if len(sys.argv) > 1:
		username = sys.argv[1]
	else:
		print("Usage: %s username" % (sys.argv[0],))
		sys.exit()

	token = util.prompt_for_user_token(username, scope, client_id = ClientID, client_secret = ClientSecret, redirect_uri = RedirectURI)


	if token:
		sp = spotipy.Spotify(auth=token)
		results = sp.current_user_saved_tracks()
		getArtistsFromPlaylist(sp)
		for i in title:
			sheet.cell(row=1, column= (title.index(i) + 1)).value = i
		artistCount = 0
		for artistIndex in uris:
		 	artistTopSongSearch(artistIndex, artistCount, sp)
		 	artistCount += 1
		wb.save('test2.xlsx')

	else:
		print("Can't get token for", username)

if __name__ == "__main__":
	main()
